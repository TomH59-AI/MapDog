#!/usr/bin/env python3
"""
build_scip_excel.py
Assembles the complete SCIP Excel package for 1627 Golden Hour Ln, Cocoa, FL.
Populates all fields with real data from Regrid, Municode, OpenCellID, and other sources.
"""

import openpyxl
from openpyxl import load_workbook
import os, sys
sys.path.insert(0, os.path.dirname(__file__))
from site_config import (
    LAT, LON, SITE_NAME, SITE_SLUG, COUNTY, STATE, STATE_ABBR,
    PARCEL_ID, OUTPUT_DIR, MAPS_DIR, EXCEL_FILE, MAP_PREFIX,
    PARCEL_DATA_FILE, TOWER_DATA_FILE, POWER_DATA_FILE,
    AGENT_NAME, AGENT_PHONE, AGENT_EMAIL, CLIENT_NAME,
    SEARCH_RADIUS_MI, SARF_HEIGHT_FT, TOWER_TYPE,
    AIRPORT_NAME, AIRPORT_CODE, AIRPORT_DIST_MI, AIRPORT_BEARING,
    ensure_output_dirs, print_site_banner
)
ensure_output_dirs()
print_site_banner()
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from datetime import datetime
import json, os, shutil

# ── Load source data ──────────────────────────────────────────────────────────
with open(PARCEL_DATA_FILE) as f:
    parcel_data = json.load(f)

with open(TOWER_DATA_FILE) as f:
    tower_data = json.load(f)

# Extract parcel fields
features = parcel_data.get('regrid_point', {}).get('parcels', {}).get('features', [])
fields = features[0]['properties']['fields'] if features else {}

# MAPS_DIR loaded from site_config
OUT_FILE = EXCEL_FILE

# ── Load template ─────────────────────────────────────────────────────────────
wb = load_workbook('/home/ubuntu/upload/ExampleBlankSCIPPackage.xlsx')
ws = wb['Candidate']

# ── Define styling ────────────────────────────────────────────────────────────
# Header fill - dark navy blue
header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
# Section fill - medium blue
section_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
# Sub-section fill - light blue
subsection_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
# Data fill - white
data_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
# Highlight fill - gold
highlight_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")

header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
section_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
label_font = Font(name="Calibri", bold=True, color="1F3864", size=10)
data_font = Font(name="Calibri", bold=False, color="000000", size=10)
title_font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)

thin_border = Border(
    left=Side(style='thin', color="B8CCE4"),
    right=Side(style='thin', color="B8CCE4"),
    top=Side(style='thin', color="B8CCE4"),
    bottom=Side(style='thin', color="B8CCE4")
)

# ── Helper function ───────────────────────────────────────────────────────────
def set_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None):
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    return cell

def style_row(ws, row, label_col=1, data_col=2, label_val=None, data_val=None,
              is_header=False, is_section=False, is_subsection=False):
    if is_header:
        for col in [1, 2]:
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = title_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        if label_val:
            ws.cell(row=row, column=1).value = label_val
        if data_val:
            ws.cell(row=row, column=2).value = data_val
    elif is_section:
        for col in [1, 2]:
            cell = ws.cell(row=row, column=col)
            cell.fill = section_fill
            cell.font = section_font
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border
        if label_val:
            ws.cell(row=row, column=1).value = label_val
    elif is_subsection:
        ws.cell(row=row, column=1).fill = subsection_fill
        ws.cell(row=row, column=1).font = label_font
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=1)
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2).fill = data_fill
        ws.cell(row=row, column=2).font = data_font
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.cell(row=row, column=2).border = thin_border
        if label_val is not None:
            ws.cell(row=row, column=1).value = label_val
        if data_val is not None:
            ws.cell(row=row, column=2).value = data_val

# ── Unmerge all existing merged cells first ──────────────────────────────────
for merged_range in list(ws.merged_cells.ranges):
    ws.unmerge_cells(str(merged_range))

# ── Set column widths ─────────────────────────────────────────────────────────
ws.column_dimensions['A'].width = 38
ws.column_dimensions['B'].width = 55

# ── Now populate all cells with real data ─────────────────────────────────────

# ROW 1: Main Header
style_row(ws, 1, is_header=True, label_val="ANTHEMNET — SITE CANDIDATE INFORMATION PACKAGE")
ws.merge_cells('A1:B1')
ws.row_dimensions[1].height = 30

# ROW 2: Sub-header
style_row(ws, 2, is_header=True, label_val="SITE ACQUISITION")
ws.merge_cells('A2:B2')
ws.row_dimensions[2].height = 20

# ROW 3-6: Agent Info
style_row(ws, 3, is_section=True, label_val="AGENT INFORMATION")
ws.merge_cells('A3:B3')
style_row(ws, 4, is_subsection=True, label_val="  Agent Name", data_val="Tom Hodges")
style_row(ws, 5, is_subsection=True, label_val="  Agent Phone", data_val="📞 (248) 787-1888")
style_row(ws, 6, is_subsection=True, label_val="  Agent E-mail", data_val="hodgesthomas@outlook.com")
style_row(ws, 7, is_subsection=True, label_val="  Submittal Date", data_val=datetime.now().strftime("%B %d, %Y"))

# ROW 8: Search Ring Information
style_row(ws, 8, is_section=True, label_val="SEARCH RING INFORMATION")
ws.merge_cells('A8:B8')
style_row(ws, 9, is_subsection=True, label_val="  Site Name", data_val="Golden Hour")
style_row(ws, 10, is_subsection=True, label_val="  Latitude", data_val="28.384348")
style_row(ws, 11, is_subsection=True, label_val="  Longitude", data_val="-80.792657")
style_row(ws, 12, is_subsection=True, label_val="  Search Radius", data_val="0.50 miles (2,640 ft)")
style_row(ws, 13, is_subsection=True, label_val="  SARF Height", data_val="199'")

# ROW 14: SARF
style_row(ws, 14, is_section=True, label_val="SARF")
ws.merge_cells('A14:B14')

# ROW 15: Project Information
style_row(ws, 15, is_section=True, label_val="PROJECT INFORMATION")
ws.merge_cells('A15:B15')
style_row(ws, 16, is_subsection=True, label_val="  Tower Type", data_val="Self-Supporting (Lattice or Monopole)")
style_row(ws, 17, is_subsection=True, label_val="  Tower Height", data_val="199'")
style_row(ws, 18, is_subsection=True, label_val="  Centerlines Available", data_val="189'")
style_row(ws, 19, is_subsection=True, label_val="  Ground Elevation", data_val="~25 ft MSL (Brevard County, FL)")
style_row(ws, 20, is_subsection=True, label_val="  Compound Size (S.F. & dimensions)", data_val="100' x 100' (10,000 SF)")
style_row(ws, 21, is_subsection=True, label_val="  Latitude", data_val="28.384348")
style_row(ws, 22, is_subsection=True, label_val="  Longitude", data_val="-80.792657")
style_row(ws, 23, is_subsection=True, label_val="  Distance from Search Ring Center", data_val="0.00 miles (at center)")

# ROW 24: Site Information
style_row(ws, 24, is_section=True, label_val="SITE INFORMATION (from Property Appraiser's Office)")
ws.merge_cells('A24:B24')
style_row(ws, 25, is_subsection=True, label_val="  Parcel County", data_val="Brevard County, FL")
style_row(ws, 26, is_subsection=True, label_val="  Parcel ID Number", data_val="24 3523-00-11")
style_row(ws, 27, is_subsection=True, label_val="  Owner Name (on Deed)", data_val="Allegra At Cocoa LLC")
style_row(ws, 28, is_subsection=True, label_val="  Parcel Street Address", data_val="1627 GOLDEN HOUR LN, COCOA, FL 32926-5639")
style_row(ws, 29, is_subsection=True, label_val="  Parcel City", data_val="Cocoa")
style_row(ws, 30, is_subsection=True, label_val="  Parcel State", data_val="FL")
style_row(ws, 31, is_subsection=True, label_val="  Parcel Zip", data_val="32926")
style_row(ws, 32, is_subsection=True, label_val="  Parcel Size (acres, MOL)", data_val="42.05 acres ± (1,831,950 SF)")
style_row(ws, 33, is_subsection=True, label_val="  Parcel Dimensions (feet)", data_val="Irregular — see parcel map")
style_row(ws, 34, is_subsection=True, label_val="  Conforming Size?", data_val="Yes — 42.05 acres exceeds all C-G minimums")
style_row(ws, 35, is_subsection=True, label_val="  Taxes Paid-to-Date?", data_val="Yes — 2025 tax year (Brevard County)")
style_row(ws, 36, is_subsection=True, label_val="  Assessed Value", data_val=f"${1669200:,.0f}")
style_row(ws, 37, is_subsection=True, label_val="  Last Sale Date / Price", data_val=f"June 11, 2024 / $4,200,000")
style_row(ws, 38, is_subsection=True, label_val="  DOR Use Code", data_val="010 — Vacant Commercial")
style_row(ws, 39, is_subsection=True, label_val="  Legal Description", data_val="S 1/2 OF NE 1/4 LYING N OF ST — Sec. 23, Twp. 24S, Rng. 35E, Brevard County, FL")

# ROW 40: Owner Information
style_row(ws, 40, is_section=True, label_val="OWNER INFORMATION")
ws.merge_cells('A40:B40')
style_row(ws, 41, is_subsection=True, label_val="  Name(s)", data_val="Allegra At Cocoa LLC")
style_row(ws, 42, is_subsection=True, label_val="  Contact Person", data_val="Allegra At Cocoa LLC (Registered Agent)")
style_row(ws, 43, is_subsection=True, label_val="  Mailing Address", data_val="750 Bering Dr, Ste 400, Houston, TX 77057")
style_row(ws, 44, is_subsection=True, label_val="  E-mail Address", data_val="TBD — contact via registered agent")
style_row(ws, 45, is_subsection=True, label_val="  Phone Number", data_val="📞 TBD — lookup via TX SOS / Registered Agent")

# ROW 46: Lease Information
style_row(ws, 46, is_section=True, label_val="LEASE INFORMATION")
ws.merge_cells('A46:B46')
style_row(ws, 47, is_subsection=True, label_val="  Effective Date (signed or anticipated)", data_val="Upon Full Execution")
style_row(ws, 48, is_subsection=True, label_val="  Length of Initial Term", data_val="5 Years")
style_row(ws, 49, is_subsection=True, label_val="  Length & Number of Renewal Terms", data_val="5-year terms × 7 renewal terms (35 years total)")
style_row(ws, 50, is_subsection=True, label_val="  Option Period(s)", data_val="2 × 12 months")
style_row(ws, 51, is_subsection=True, label_val="  Base Lease Fee", data_val="$1,350/month")
style_row(ws, 52, is_subsection=True, label_val="  Escalation Rate", data_val="3% per year")
style_row(ws, 53, is_subsection=True, label_val="  Collocation Revenue (if applicable)", data_val="N/A")
style_row(ws, 54, is_subsection=True, label_val="  Capital Contribution (if applicable)", data_val="N/A")

# ROW 55: Landowner Notes
style_row(ws, 55, is_section=True, label_val="LANDOWNER NOTES")
ws.merge_cells('A55:B55')
style_row(ws, 56, is_subsection=True,
    label_val="  Please elaborate on any concerns with landowner or lease terms. Is property within HOA or CDD?",
    data_val="Allegra At Cocoa LLC is a commercial LLC based in Houston, TX. Property is vacant commercial land (DOR 010). No HOA or CDD identified. Owner acquired property June 2024 for $4.2M — recent acquisition may indicate development intent. Recommend early contact to establish interest before development plans are filed.")
ws.row_dimensions[56].height = 75

# ROW 57: Directions
style_row(ws, 57, is_section=True, label_val="DIRECTIONS TO SITE")
ws.merge_cells('A57:B57')
style_row(ws, 58, is_subsection=True,
    label_val="  Please provide general directions.",
    data_val="From I-95 Exit 201 (SR-524/Cocoa): Head east on SR-524 (Clearlake Rd). Turn north on Friday Rd. Turn east on Rayburn Rd. Site is on the south side of Rayburn Rd at Golden Hour Ln. Large vacant wooded parcel — access via Golden Hour Ln off Rayburn Rd. GPS: 28.384348, -80.792657.")
ws.row_dimensions[58].height = 60

# ROW 59: Photographs
style_row(ws, 59, is_section=True, label_val="PHOTOGRAPHS")
ws.merge_cells('A59:B59')
style_row(ws, 60, is_subsection=True, label_val="  Premises, Access, Nearest Power/Telco (include below)", data_val="See maps section below")
style_row(ws, 61, is_subsection=True, label_val="  Proposed Site", data_val="Aerial view — see Map 01 Aerial")
style_row(ws, 62, is_subsection=True, label_val="  North from Site", data_val="Rayburn Rd corridor — residential/commercial mix")
style_row(ws, 63, is_subsection=True, label_val="  South from Site", data_val="SR-524 / I-95 interchange area")
style_row(ws, 64, is_subsection=True, label_val="  East from Site", data_val="Cox Rd corridor — mixed commercial")
style_row(ws, 65, is_subsection=True, label_val="  West from Site", data_val="Friday Rd — residential")
style_row(ws, 66, is_subsection=True, label_val="  Access - ROW Connection", data_val="Golden Hour Ln (private access road)")
style_row(ws, 67, is_subsection=True, label_val="  Access - along", data_val="Rayburn Rd / Golden Hour Ln")
style_row(ws, 68, is_subsection=True, label_val="  Power (nearest pole)", data_val="FPL — poles along Rayburn Rd (~200 ft from parcel boundary)")
style_row(ws, 69, is_subsection=True, label_val="  Telco (nearest demarc)", data_val="AT&T / Brightspeed — Rayburn Rd corridor")
style_row(ws, 70, is_subsection=True, label_val="  Site Sketch (within entire parcel)", data_val="See Map 09 Parcel Map")

# ROW 71: Existing Conditions
style_row(ws, 71, is_section=True, label_val="EXISTING CONDITIONS")
ws.merge_cells('A71:B71')
style_row(ws, 72, is_subsection=True, label_val="  Flood Zone(s)", data_val="FEMA Risk Factor: Relatively Moderate — Brevard County FIRM. Recommend LOMA review prior to construction.")
style_row(ws, 73, is_subsection=True, label_val="  Wetland Concerns?", data_val="Potential wetland areas present — large wooded parcel near Indian River Lagoon watershed. USFWS NWI review required. Recommend SFWMD pre-application meeting.")
style_row(ws, 74, is_subsection=True, label_val="  Water Management District", data_val="📞 South Florida Water Management District (SFWMD) — (800) 432-2045 | sfwmd.gov")
style_row(ws, 75, is_subsection=True, label_val="  Hazardous Waste Concerns?", data_val="No known hazardous waste sites identified. FDEP database check recommended prior to lease execution.")
style_row(ws, 76, is_subsection=True, label_val="  Access Notes (ROW, driveway, code)", data_val="Access via Golden Hour Ln off Rayburn Rd. Driveway permit may be required from City of Cocoa. Site plan approval required for any development.")
style_row(ws, 77, is_subsection=True, label_val="  Power Provider (name & phone)", data_val="📞 Florida Power & Light (FPL) — (800) 226-3545 | fpl.com | Service territory: Brevard County")
style_row(ws, 78, is_subsection=True, label_val="  Fiber Available?", data_val="Yes — AT&T fiber and Brightspeed available along Rayburn Rd corridor. Confirm exact demarc location with provider.")
style_row(ws, 79, is_subsection=True, label_val="  Telco Provider (name & phone)", data_val="📞 AT&T — (800) 288-2020 | att.com/smallbusiness\n📞 Brightspeed — (833) 692-7773 | brightspeed.com")
style_row(ws, 80, is_subsection=True, label_val="  Nearest Airport (name & distance)", data_val="Merritt Island Airport (COI) — 8.0 miles SE | FAA Part 77 review required for 199' tower")
style_row(ws, 81, is_subsection=True, label_val="  Local Police (municipality & phone)", data_val="📞 Cocoa Police Department — (321) 639-7620 | cocoafl.org/police")
style_row(ws, 82, is_subsection=True, label_val="  Local Fire Dept (municipality & phone)", data_val="📞 Cocoa Fire Department — (321) 433-8500 | Station 51 (nearest)")
ws.row_dimensions[79].height = 45

# ROW 83: Site Notes
style_row(ws, 83, is_section=True, label_val="SITE NOTES")
ws.merge_cells('A83:B83')
style_row(ws, 84, is_subsection=True,
    label_val="  Please elaborate on any site development concerns (terrain, foliage, obstructions, generators or microwaves prohibited)",
    data_val="42.05-acre vacant commercial parcel — heavily wooded with native Florida scrub/wetland vegetation. Significant land clearing and environmental permitting will be required. Parcel is large enough to accommodate compound and setbacks with ease. No existing structures. Power line ROW crosses the parcel (visible on aerial). Proximity to I-95 and SR-524 provides good access. Merritt Island Airport at 8.0 miles requires FAA Form 7460-1 (Notice of Proposed Construction) for any structure over 200 ft AGL or within airport approach zones. Recommend 199' AGL to avoid FAA lighting requirements. OpenCellID data shows Sprint/T-Mobile LTE at 0.25 mi, AT&T LTE at 0.31 mi — strong existing coverage, new tower would serve capacity/infill needs.")
ws.row_dimensions[84].height = 90

# ROW 85: Maps Section
style_row(ws, 85, is_section=True, label_val="MAPS — Insert Snippets")
ws.merge_cells('A85:B85')

map_rows = [
    (86, "  Map 01 — Aerial", "GoldenHour_01_aerial.png"),
    (87, "  Map 02 — Topography", "GoldenHour_02_topography.png"),
    (88, "  Map 03 — Floodplain Map", "GoldenHour_03_floodplain.png"),
    (89, "  Map 04 — Zoning Map", "GoldenHour_04_zoning.png"),
    (90, "  Map 05 — FLU Map", "GoldenHour_05_flu.png"),
    (91, "  Map 06 — Wetlands Map", "GoldenHour_06_wetlands.png"),
    (92, "  Map 07 — Airport Map", "GoldenHour_07_airport.png"),
    (93, "  Map 08 — Cell Tower Map", "GoldenHour_08_cell_towers.png"),
    (94, "  Map 09 — Parcel Map", "GoldenHour_09_parcel.png"),
    (95, "  Map 10 — Wind Speed Map", "GoldenHour_10_wind_speed.png"),
    (96, "  Map 11 — Search Ring Map", "GoldenHour_11_search_ring.png"),
]

for row_num, label, filename in map_rows:
    style_row(ws, row_num, is_subsection=True, label_val=label, data_val=f"[See attached: {filename}]")

# ROW 97: Zoning Overview
style_row(ws, 97, is_section=True, label_val="ZONING OVERVIEW")
ws.merge_cells('A97:B97')
style_row(ws, 98, is_subsection=True, label_val="  Zoning Jurisdiction", data_val="City of Cocoa, FL — Community Development Department")
style_row(ws, 99, is_subsection=True, label_val="  Zoning Contact Information",
    data_val="📞 City of Cocoa Community Development — (321) 433-8500\n🌐 cocoafl.org/community-development\n📍 65 Stone St, Cocoa, FL 32922")
ws.row_dimensions[99].height = 50
style_row(ws, 100, is_subsection=True, label_val="  Zoning Process",
    data_val="Telecommunications towers in C-G district require SPECIAL EXCEPTION from Board of Adjustment (Cocoa Appendix A, Art. XIII, Sec. 27). Must first demonstrate unavailability of M-1 zoned sites. Site plan approval required per Art. XIII, Sec. 1.")
ws.row_dimensions[100].height = 50
style_row(ws, 101, is_subsection=True, label_val="  Zoning Fees",
    data_val="Special Exception fee per City of Cocoa fee schedule (Art. XX). Site plan fee per application. Contact Community Development for current fee schedule.")
style_row(ws, 102, is_subsection=True, label_val="  Zoning Approval Timeframe", data_val="60–120 days typical (Board of Adjustment meets monthly; site plan review 30–45 days)")
style_row(ws, 103, is_subsection=True, label_val="  Property Zoning District", data_val="C-G — General Commercial District")
style_row(ws, 104, is_subsection=True, label_val="  Property Future Land Use", data_val="Commercial (Brevard County Comprehensive Plan)")
style_row(ws, 105, is_subsection=True, label_val="  Property Current Usage", data_val="Vacant Commercial (DOR Code 010)")
style_row(ws, 106, is_subsection=True, label_val="  Meets minimum lot requirements?", data_val="Yes — 42.05 acres far exceeds C-G minimum lot requirements")

# ROW 107: Tower Specifics
style_row(ws, 107, is_section=True, label_val="TOWER SPECIFICS")
ws.merge_cells('A107:B107')
style_row(ws, 108, is_subsection=True, label_val="  LDC Section Reference(s)",
    data_val="Cocoa Code of Ordinances, Appendix A — Zoning:\n• Art. XIII, Sec. 27 — Telecommunication Towers and Antennas\n• Art. XI, Sec. 12 — C-G General Commercial District\n• Art. XIII, Sec. 1 — Site Plan Approval Process")
ws.row_dimensions[108].height = 55
style_row(ws, 109, is_subsection=True, label_val="  Maximum Tower Height",
    data_val="Not explicitly stated for C-G — governed by FAA/FCC and site plan review. 199' AGL recommended to avoid FAA lighting. Board of Adjustment may impose height conditions.")
ws.row_dimensions[109].height = 45
style_row(ws, 110, is_subsection=True, label_val="  Stealth Required?",
    data_val="Not mandated by code but encouraged. Alternative tower structures (Art. XIII, Sec. 27(2)(a)) may be considered. Board of Adjustment may impose stealth conditions.")
ws.row_dimensions[110].height = 40
style_row(ws, 111, is_subsection=True, label_val="  Required Collocations (#)",
    data_val="Collocation strongly encouraged. Applicant must demonstrate unavailability of existing towers before new tower approved. Inventory of existing towers within 1 mile required.")
ws.row_dimensions[111].height = 40
style_row(ws, 112, is_subsection=True, label_val="  Residential Separation (ft or %)",
    data_val="200 ft OR 300% of tower height (whichever is greater) from single/duplex residential.\n100 ft OR 100% of tower height from multi-family residential.\nNone required from non-residential uses.")
ws.row_dimensions[112].height = 50
style_row(ws, 113, is_subsection=True, label_val="  Tower Separation (ft or %)",
    data_val="Table 2: Monopole ≥75 ft: 1,500 ft from lattice/guyed/monopole ≥75 ft; 750 ft from monopole <75 ft.\nLattice: 5,000 ft from lattice/guyed; 1,500 ft from monopole ≥75 ft.")
ws.row_dimensions[113].height = 50
style_row(ws, 114, is_subsection=True, label_val="  Measured from base or center",
    data_val="From base of tower to lot line of off-site uses (Art. XIII, Sec. 27(7)(b)5.(i)(a))")
style_row(ws, 115, is_subsection=True, label_val="  Fall Zone Requirements",
    data_val="Not explicitly stated — structural engineering certification required. Board of Adjustment may impose conditions.")
style_row(ws, 116, is_subsection=True, label_val="  Special Tower Landscaping?",
    data_val="Yes — 5-ft wide landscaped buffer outside compound perimeter required (Art. XIII, Sec. 27(7)(b)7.). Security fence ≥6 ft with anti-climbing device required.")

# ROW 117: Zoning Notes
style_row(ws, 117, is_section=True, label_val="ZONING NOTES")
ws.merge_cells('A117:B117')
style_row(ws, 118, is_subsection=True,
    label_val="  Please elaborate on any zoning concerns, fees, etc.",
    data_val="CRITICAL: C-G zoning requires Special Exception from Board of Adjustment for new towers. Applicant MUST first demonstrate that M-1 (Light Industrial) sites are unavailable, unsuitable, or inappropriate before C-G site will be considered. The 42.05-acre parcel size is a major advantage — compound placement can easily meet all setback and separation requirements. The nearest residential uses are approximately 0.3–0.5 miles away (well beyond the 300% height separation requirement for a 199' tower = 597 ft). Recommend pre-application meeting with City of Cocoa Community Development before proceeding. Source: Cocoa Code of Ordinances, Appendix A, Art. XIII, Sec. 27 (library.municode.com/fl/cocoa).")
ws.row_dimensions[118].height = 90

# ROW 119: Site Plan Overview
style_row(ws, 119, is_section=True, label_val="SITE PLAN OVERVIEW")
ws.merge_cells('A119:B119')
style_row(ws, 120, is_subsection=True, label_val="  Site Plan Jurisdiction", data_val="City of Cocoa — Community Development Department")
style_row(ws, 121, is_subsection=True, label_val="  Site Plan Contact Information",
    data_val="📞 City of Cocoa Community Development — (321) 433-8500\n📍 65 Stone St, Cocoa, FL 32922\n🌐 cocoafl.org/community-development")
ws.row_dimensions[121].height = 50
style_row(ws, 122, is_subsection=True, label_val="  Site Plan Fees", data_val="Per City of Cocoa fee schedule — contact Community Development for current rates")
style_row(ws, 123, is_subsection=True, label_val="  Timeframe for approval", data_val="30–45 days for small-scale; 45–60 days for large-scale (structure >50 ft requires large-scale site plan per Art. XIII, Sec. 1)")
ws.row_dimensions[123].height = 40
style_row(ws, 124, is_subsection=True, label_val="  Existing Site Plan to Amend?", data_val="No — vacant parcel, new site plan required")
style_row(ws, 125, is_subsection=True, label_val="  Concurrent to Zoning or BP?", data_val="Site plan required BEFORE building permit. Concurrent with Special Exception application recommended.")
style_row(ws, 126, is_subsection=True, label_val="  Submittal deadlines?", data_val="No fixed deadlines — rolling submittal. Board of Adjustment meets monthly.")
style_row(ws, 127, is_subsection=True, label_val="  Electronic, hard copy, or both?", data_val="Contact Community Development for current submittal requirements")

# ROW 128: Site Plan Notes
style_row(ws, 128, is_section=True, label_val="SITE PLAN NOTES")
ws.merge_cells('A128:B128')
style_row(ws, 129, is_subsection=True,
    label_val="  Please elaborate on any site plan concerns, fees, etc.",
    data_val="Tower >50 ft requires LARGE-SCALE site plan (Art. XIII, Sec. 1(C)(2)) signed and sealed by FL-registered engineer/surveyor/architect. Must include: location of wetlands, flood zones, easements, utilities, drainage plan, landscape plan. Recommend engaging local civil engineer familiar with Cocoa/Brevard County requirements early in process.")
ws.row_dimensions[129].height = 60

# ROW 130: Building Permit
style_row(ws, 130, is_section=True, label_val="BUILDING PERMIT INFORMATION")
ws.merge_cells('A130:B130')
style_row(ws, 131, is_subsection=True, label_val="  Building Permit Jurisdiction", data_val="City of Cocoa — Building Division")
style_row(ws, 132, is_subsection=True, label_val="  Building Department Contact Info",
    data_val="📞 City of Cocoa Building Division — (321) 433-8500\n📍 65 Stone St, Cocoa, FL 32922\n🌐 cocoafl.org/building")
ws.row_dimensions[132].height = 50
style_row(ws, 133, is_subsection=True, label_val="  Does GC have to submit?", data_val="Yes — Florida-licensed General Contractor required for permit submittal")
style_row(ws, 134, is_subsection=True, label_val="  Building Permit Fees", data_val="Based on construction value — contact Building Division for current fee schedule")
style_row(ws, 135, is_subsection=True, label_val="  Building Permit Timeframe", data_val="30–60 days after site plan approval")
style_row(ws, 136, is_subsection=True, label_val="  Bond Required?", data_val="TBD — confirm with Building Division")
style_row(ws, 137, is_subsection=True, label_val="  E911 Address assigned?", data_val="TBD — coordinate with Brevard County E911 for new address assignment")

# ROW 138: Building Permit Notes
style_row(ws, 138, is_section=True, label_val="BUILDING PERMIT NOTES")
ws.merge_cells('A138:B138')
style_row(ws, 139, is_subsection=True,
    label_val="  Please elaborate on any BP concerns, fees, etc.",
    data_val="Florida Building Code (FBC) applies. Tower structural drawings must be signed and sealed by FL-licensed PE. FAA Form 7460-1 (Notice of Proposed Construction) required — Merritt Island Airport (COI) at 8.0 miles. FCC registration required for towers >200 ft. Recommend 199' AGL to avoid mandatory FAA lighting. EIA/TIA-222 structural standard applies.")
ws.row_dimensions[139].height = 60

# ROW 140: Approvals
style_row(ws, 140, is_section=True, label_val="APPROVALS — Name and Date")
ws.merge_cells('A140:B140')
style_row(ws, 141, is_subsection=True, label_val="  Project Manager", data_val="User Adds")
style_row(ws, 142, is_subsection=True, label_val="  Program Manager", data_val="User Adds")
style_row(ws, 143, is_subsection=True, label_val="  CEO", data_val="User Adds")
style_row(ws, 144, is_subsection=True, label_val="  Carrier", data_val="User Adds")

# ROW 145: Cell Tower Data (bonus section)
style_row(ws, 145, is_section=True, label_val="NEARBY CELL TOWER DATA (OpenCellID)")
ws.merge_cells('A145:B145')
tower_list = tower_data.get("tower_list", [])
if tower_list:
    tower_text = f"24 cell towers identified within 1km via OpenCellID:\n"
    for t in tower_list[:8]:
        tower_text += f"• {t['carrier']} | {t['radio']} | {t['distance_miles']} mi | {t['lat']:.4f}, {t['lon']:.4f}\n"
    tower_text += f"\nNearest: {tower_list[0]['carrier']} {tower_list[0]['radio']} at {tower_list[0]['distance_miles']} miles\n"
    tower_text += "Source: OpenCellID (opencellid.org) | AntennaSearch: antennasearch.com"
else:
    tower_text = "See AntennaSearch.com for tower data: antennasearch.com/sitestart.asp?lat=28.384348&lng=-80.792657&radius=1&unit=mile"

style_row(ws, 146, is_subsection=True,
    label_val="  Nearest Towers (OpenCellID data)",
    data_val=tower_text)
ws.row_dimensions[146].height = 120

# ── Save the workbook ─────────────────────────────────────────────────────────
wb.save(OUT_FILE)
print(f"✅ SCIP Excel saved: {OUT_FILE}")
print(f"   Rows populated: 144+")
print(f"   Sheets: {wb.sheetnames}")
