#!/usr/bin/env python3
"""
embed_maps.py
Embeds all 11 SCIP maps as images into the Excel package.
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from PIL import Image as PILImage
import os, io, sys
sys.path.insert(0, os.path.dirname(__file__))
from site_config import (
    SITE_NAME, SITE_SLUG, OUTPUT_DIR, MAPS_DIR, EXCEL_FILE, MAP_PREFIX,
    ensure_output_dirs, print_site_banner
)
ensure_output_dirs()
print_site_banner()

# MAPS_DIR from site_config
EXCEL_FILE = EXCEL_FILE  # from site_config

wb = load_workbook(EXCEL_FILE)
ws = wb['Candidate']

# Map definitions: (label, filename, row_to_insert_after)
MAPS = [
    ("Map 01 — Aerial View", f"{MAP_PREFIX}_01_aerial.png"),
    ("Map 02 — Topography", f"{MAP_PREFIX}_02_topography.png"),
    ("Map 03 — FEMA Floodplain", f"{MAP_PREFIX}_03_floodplain.png"),
    ("Map 04 — Zoning", f"{MAP_PREFIX}_04_zoning.png"),
    ("Map 05 — Future Land Use", f"{MAP_PREFIX}_05_flu.png"),
    ("Map 06 — USFWS Wetlands", f"{MAP_PREFIX}_06_wetlands.png"),
    ("Map 07 — Airport Proximity", f"{MAP_PREFIX}_07_airport.png"),
    ("Map 08 — Cell Towers", f"{MAP_PREFIX}_08_cell_towers.png"),
    ("Map 09 — Parcel Map", f"{MAP_PREFIX}_09_parcel.png"),
    ("Map 10 — Wind Speed", "GoldenHour_10_wind_speed.png"),
    ("Map 11 — Search Ring", f"{MAP_PREFIX}_11_search_ring.png"),
]

# Create a new "Maps" sheet
if "Maps" in wb.sheetnames:
    del wb["Maps"]
ws_maps = wb.create_sheet("Maps")

# Style the maps sheet
header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
section_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
section_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)

ws_maps.column_dimensions['A'].width = 3
ws_maps.column_dimensions['B'].width = 80
ws_maps.column_dimensions['C'].width = 3

# Header
ws_maps.merge_cells('A1:C1')
ws_maps['A1'] = "GOLDEN HOUR — SITE CANDIDATE INFORMATION PACKAGE — MAP SET"
ws_maps['A1'].fill = header_fill
ws_maps['A1'].font = header_font
ws_maps['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_maps.row_dimensions[1].height = 35

ws_maps.merge_cells('A2:C2')
ws_maps['A2'] = "1627 Golden Hour Ln, Cocoa, FL 32926  |  Lat: 28.384348  |  Lon: -80.792657  |  Parcel: 24 3523-00-11"
ws_maps['A2'].fill = section_fill
ws_maps['A2'].font = section_font
ws_maps['A2'].alignment = Alignment(horizontal='center', vertical='center')
ws_maps.row_dimensions[2].height = 22

current_row = 3

for i, (label, filename) in enumerate(MAPS, 1):
    map_path = os.path.join(MAPS_DIR, filename)
    if not os.path.exists(map_path):
        print(f"  ⚠️  Missing: {filename}")
        continue
    
    # Add label row
    ws_maps.merge_cells(f'A{current_row}:C{current_row}')
    ws_maps[f'A{current_row}'] = f"  {i:02d}. {label}"
    ws_maps[f'A{current_row}'].fill = section_fill
    ws_maps[f'A{current_row}'].font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    ws_maps[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws_maps.row_dimensions[current_row].height = 22
    current_row += 1
    
    # Resize image to fit nicely in Excel (max 900px wide, maintain aspect ratio)
    try:
        with PILImage.open(map_path) as pil_img:
            orig_w, orig_h = pil_img.size
            max_w = 900
            if orig_w > max_w:
                ratio = max_w / orig_w
                new_w = max_w
                new_h = int(orig_h * ratio)
            else:
                new_w, new_h = orig_w, orig_h
            
            # Resize and save to buffer
            resized = pil_img.resize((new_w, new_h), PILImage.LANCZOS)
            buf = io.BytesIO()
            resized.save(buf, format='PNG')
            buf.seek(0)
        
        # Calculate row height needed (Excel row height in points, 1 pt ≈ 0.75 px)
        img_height_pts = new_h * 0.75
        # Set rows for the image (need enough rows)
        rows_needed = max(30, int(img_height_pts / 15) + 2)
        
        # Set row heights
        for r in range(current_row, current_row + rows_needed):
            ws_maps.row_dimensions[r].height = 15
        
        # Add image
        xl_img = XLImage(buf)
        xl_img.width = new_w * 0.75  # Convert pixels to points
        xl_img.height = new_h * 0.75
        xl_img.anchor = f'B{current_row}'
        ws_maps.add_image(xl_img)
        
        current_row += rows_needed + 2
        print(f"  ✅ Embedded: {filename} ({orig_w}x{orig_h} → {new_w}x{new_h})")
    
    except Exception as e:
        print(f"  ❌ Error embedding {filename}: {e}")
        current_row += 2

# Also embed maps inline in the Candidate sheet at the map rows
print("\nEmbedding thumbnail maps in Candidate sheet...")
map_row_mapping = {
    86: f"{MAP_PREFIX}_01_aerial.png",
    87: f"{MAP_PREFIX}_02_topography.png",
    88: f"{MAP_PREFIX}_03_floodplain.png",
    89: f"{MAP_PREFIX}_04_zoning.png",
    90: f"{MAP_PREFIX}_05_flu.png",
    91: f"{MAP_PREFIX}_06_wetlands.png",
    92: f"{MAP_PREFIX}_07_airport.png",
    93: f"{MAP_PREFIX}_08_cell_towers.png",
    94: f"{MAP_PREFIX}_09_parcel.png",
    95: "GoldenHour_10_wind_speed.png",
    96: f"{MAP_PREFIX}_11_search_ring.png",
}

for row_num, filename in map_row_mapping.items():
    map_path = os.path.join(MAPS_DIR, filename)
    if not os.path.exists(map_path):
        continue
    try:
        with PILImage.open(map_path) as pil_img:
            # Thumbnail size for inline view
            thumb_w, thumb_h = 400, 300
            resized = pil_img.resize((thumb_w, thumb_h), PILImage.LANCZOS)
            buf = io.BytesIO()
            resized.save(buf, format='PNG')
            buf.seek(0)
        
        xl_img = XLImage(buf)
        xl_img.width = 300  # points
        xl_img.height = 225  # points
        xl_img.anchor = f'B{row_num}'
        ws.add_image(xl_img)
        ws.row_dimensions[row_num].height = 170
        print(f"  ✅ Thumbnail in row {row_num}: {filename}")
    except Exception as e:
        print(f"  ❌ Row {row_num} error: {e}")

wb.save(EXCEL_FILE)
print(f"\n✅ Final SCIP package saved with maps: {EXCEL_FILE}")
print(f"   File size: {os.path.getsize(EXCEL_FILE):,} bytes")
